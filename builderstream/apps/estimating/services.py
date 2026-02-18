"""Estimating services for calculations, proposals, and exports."""
import base64
import logging
from decimal import Decimal
from datetime import timedelta
from io import BytesIO

from django.conf import settings
from django.core.files.base import ContentFile
from django.db import transaction, models
from django.utils import timezone

from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from apps.projects.models import ActivityLog

logger = logging.getLogger(__name__)


class EstimateCalculationService:
    """Service for estimate calculations and totals."""

    @staticmethod
    def calculate_estimate_totals(estimate):
        """Recalculate all estimate totals from line items."""
        from .models import EstimateLineItem

        subtotal = Decimal("0.00")

        for section in estimate.sections.all():
            section_total = Decimal("0.00")
            for line_item in section.line_items.all():
                # Calculate line total
                line_item.line_total = line_item.quantity * line_item.unit_price
                line_item.save(update_fields=["line_total"])

                # Add to section total if taxable
                if line_item.is_taxable:
                    section_total += line_item.line_total

            section.subtotal = section_total
            section.save(update_fields=["subtotal"])
            subtotal += section_total

        # Calculate tax and total
        estimate.subtotal = subtotal
        estimate.tax_amount = (subtotal * estimate.tax_rate) / Decimal("100")
        estimate.total = subtotal + estimate.tax_amount
        estimate.save(update_fields=["subtotal", "tax_amount", "total"])

        return estimate

    @staticmethod
    def calculate_assembly_totals(assembly):
        """Recalculate assembly totals from assembly items."""
        total_cost = Decimal("0.00")
        total_price = Decimal("0.00")

        for item in assembly.assembly_items.all():
            total_cost += item.cost_item.cost * item.quantity
            total_price += item.cost_item.client_price * item.quantity

        assembly.total_cost = total_cost
        assembly.total_price = total_price
        assembly.save(update_fields=["total_cost", "total_price"])

        return assembly

    @staticmethod
    def copy_estimate(estimate, user, new_name=None):
        """Create a copy of an estimate with all sections and line items."""
        from .models import Estimate, EstimateSection, EstimateLineItem

        with transaction.atomic():
            # Create new estimate
            new_estimate = Estimate.objects.create(
                organization=estimate.organization,
                project=estimate.project,
                lead=estimate.lead,
                name=new_name or f"{estimate.name} (Copy)",
                status="draft",
                tax_rate=estimate.tax_rate,
                notes=estimate.notes,
                valid_until=estimate.valid_until,
                created_by=user,
            )

            # Copy all sections
            for section in estimate.sections.all():
                new_section = EstimateSection.objects.create(
                    organization=estimate.organization,
                    estimate=new_estimate,
                    name=section.name,
                    description=section.description,
                    sort_order=section.sort_order,
                )

                # Copy all line items
                for line_item in section.line_items.all():
                    EstimateLineItem.objects.create(
                        organization=estimate.organization,
                        section=new_section,
                        cost_item=line_item.cost_item,
                        assembly=line_item.assembly,
                        description=line_item.description,
                        quantity=line_item.quantity,
                        unit=line_item.unit,
                        unit_cost=line_item.unit_cost,
                        unit_price=line_item.unit_price,
                        is_taxable=line_item.is_taxable,
                        sort_order=line_item.sort_order,
                        notes=line_item.notes,
                    )

            # Recalculate totals
            EstimateCalculationService.calculate_estimate_totals(new_estimate)

            return new_estimate


class ProposalService:
    """Service for proposal generation and management."""

    @staticmethod
    def generate_proposal_from_estimate(estimate, user, client, template=None):
        """Create proposal from estimate with PDF generation."""
        from .models import Proposal, ProposalTemplate
        from .tasks import generate_pdf_proposal

        # Get default template if not provided
        if not template:
            template = ProposalTemplate.objects.filter(
                organization=estimate.organization,
                is_default=True,
            ).first()

        # Auto-generate proposal number
        org = estimate.organization
        year = timezone.now().year
        count = Proposal.objects.filter(organization=org).count()
        proposal_number = f"PROP-{year}-{count + 1:03d}"

        # Create proposal
        proposal = Proposal.objects.create(
            organization=org,
            estimate=estimate,
            project=estimate.project,
            lead=estimate.lead,
            client=client,
            template=template,
            proposal_number=proposal_number,
            status="draft",
            valid_until=estimate.valid_until,
            terms_and_conditions=template.terms_and_conditions if template else "",
        )

        # Trigger async PDF generation
        generate_pdf_proposal.delay(str(proposal.pk))

        # Log activity
        ActivityLog.objects.create(
            organization=estimate.organization,
            user=user,
            action="created",
            entity_type="proposal",
            entity_id=proposal.pk,
            description=f"Generated proposal {proposal.proposal_number} from estimate {estimate.estimate_number}",
            metadata={
                "estimate_id": str(estimate.id),
                "proposal_id": str(proposal.id),
            },
        )

        return proposal

    @staticmethod
    def send_proposal(proposal, user, recipient_email=None):
        """Send proposal via email with public link."""
        from .tasks import send_proposal_email

        if not proposal.pdf_file:
            raise ValueError("Proposal PDF not generated yet")

        if not recipient_email:
            recipient_email = proposal.client.email

        # Update proposal status
        proposal.sent_at = timezone.now()
        proposal.sent_to_email = recipient_email
        proposal.status = "sent"
        proposal.save(update_fields=["sent_at", "sent_to_email", "status"])

        # Trigger async email sending
        send_proposal_email.delay(str(proposal.pk), recipient_email)

        # Log activity
        ActivityLog.objects.create(
            organization=proposal.organization,
            user=user,
            action="sent",
            entity_type="proposal",
            entity_id=proposal.pk,
            description=f"Sent proposal {proposal.proposal_number} to {recipient_email}",
            metadata={
                "recipient": recipient_email,
                "public_token": str(proposal.public_token),
            },
        )

        return proposal

    @staticmethod
    def capture_signature(proposal, signature_data, signed_by_name, ip_address=None, user_agent=None):
        """Capture e-signature with metadata."""
        # signature_data is base64 encoded image from canvas
        # Format: "data:image/png;base64,iVBORw0KG..."
        try:
            format_part, imgstr = signature_data.split(";base64,")
            ext = format_part.split("/")[-1]  # png, jpeg, etc.
        except (ValueError, IndexError):
            raise ValueError("Invalid signature data format")

        # Decode base64 to file
        signature_file = ContentFile(
            base64.b64decode(imgstr),
            name=f"signature_{proposal.pk}.{ext}",
        )

        # Update proposal
        proposal.signature_image = signature_file
        proposal.signed_at = timezone.now()
        proposal.signed_by_name = signed_by_name
        proposal.signature_ip = ip_address
        proposal.signature_user_agent = user_agent
        proposal.is_signed = True
        proposal.status = "signed"
        proposal.save()

        # Log activity (user is None for unauthenticated signature)
        ActivityLog.objects.create(
            organization=proposal.organization,
            user=None,
            action="created",
            entity_type="proposal",
            entity_id=proposal.pk,
            description=f"Proposal {proposal.proposal_number} signed by {signed_by_name}",
            metadata={
                "signed_at": proposal.signed_at.isoformat(),
                "ip": ip_address,
                "user_agent": user_agent,
            },
        )

        return proposal


class ExportService:
    """Service for Excel and PDF exports."""

    @staticmethod
    def export_estimate_to_excel(estimate):
        """Export estimate to Excel with formulas."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Estimate"

        # Header styling
        header_font = Font(size=16, bold=True)
        section_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        total_font = Font(bold=True, size=14)

        # Organization header
        ws["A1"] = estimate.organization.name
        ws["A1"].font = header_font
        ws["A2"] = f"Estimate: {estimate.estimate_number}"
        ws["A3"] = f"Date: {estimate.created_at.strftime('%Y-%m-%d')}"

        # Column headers
        row = 5
        headers = ["Section", "Item", "Description", "Quantity", "Unit", "Unit Price", "Total", "Taxable"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = section_font
            cell.fill = header_fill

        row += 1

        # Line items
        for section in estimate.sections.all():
            # Section header
            ws.cell(row=row, column=1, value=section.name).font = section_font
            row += 1

            for item in section.line_items.all():
                ws.cell(row=row, column=1, value="")
                ws.cell(row=row, column=2, value=item.cost_item.name if item.cost_item else item.assembly.name)
                ws.cell(row=row, column=3, value=item.description or "")
                ws.cell(row=row, column=4, value=float(item.quantity))
                ws.cell(row=row, column=5, value=item.unit)
                ws.cell(row=row, column=6, value=float(item.unit_price))
                ws.cell(row=row, column=7, value=f"=D{row}*F{row}")  # Formula
                ws.cell(row=row, column=8, value="Yes" if item.is_taxable else "No")
                row += 1

            # Section subtotal
            ws.cell(row=row, column=6, value="Section Total:").font = section_font
            ws.cell(row=row, column=7, value=float(section.subtotal)).font = section_font
            row += 2

        # Grand totals
        ws.cell(row=row, column=6, value="Subtotal:").font = section_font
        ws.cell(row=row, column=7, value=float(estimate.subtotal)).font = section_font
        row += 1

        ws.cell(row=row, column=6, value=f"Tax ({estimate.tax_rate}%):").font = section_font
        ws.cell(row=row, column=7, value=float(estimate.tax_amount)).font = section_font
        row += 1

        ws.cell(row=row, column=6, value="TOTAL:").font = total_font
        ws.cell(row=row, column=7, value=float(estimate.total)).font = total_font

        # Column widths
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 40
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 12
        ws.column_dimensions["H"].width = 10

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    @staticmethod
    def generate_proposal_pdf(proposal):
        """Generate PDF from proposal using reportlab."""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
        story = []
        styles = getSampleStyleSheet()

        # Custom styles
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=24,
            textColor=colors.HexColor("#2563eb"),
            spaceAfter=30,
            alignment=TA_CENTER,
        )

        heading_style = ParagraphStyle(
            "CustomHeading",
            parent=styles["Heading2"],
            fontSize=14,
            textColor=colors.HexColor("#1e40af"),
            spaceBefore=12,
            spaceAfter=6,
        )

        # Header
        if proposal.template and proposal.template.header_text:
            story.append(Paragraph(proposal.template.header_text, styles["Normal"]))
            story.append(Spacer(1, 0.2*inch))

        story.append(Paragraph(f"PROPOSAL {proposal.proposal_number}", title_style))
        story.append(Spacer(1, 0.3*inch))

        # Client info
        client_info = f"""
        <b>Prepared For:</b><br/>
        {proposal.client.first_name} {proposal.client.last_name}<br/>
        {proposal.client.email}<br/>
        {proposal.client.phone or ''}<br/>
        """
        story.append(Paragraph(client_info, styles["Normal"]))
        story.append(Spacer(1, 0.2*inch))

        # Date and validity
        date_info = f"""
        <b>Date:</b> {proposal.created_at.strftime('%B %d, %Y')}<br/>
        <b>Valid Until:</b> {proposal.valid_until.strftime('%B %d, %Y') if proposal.valid_until else 'N/A'}<br/>
        """
        story.append(Paragraph(date_info, styles["Normal"]))
        story.append(Spacer(1, 0.3*inch))

        # Line items table
        estimate = proposal.estimate
        table_data = [["Item", "Description", "Qty", "Unit", "Price", "Total"]]

        for section in estimate.sections.all():
            # Section header row
            table_data.append([
                Paragraph(f"<b>{section.name}</b>", styles["Normal"]),
                "", "", "", "", ""
            ])

            for item in section.line_items.all():
                item_name = item.description or (item.cost_item.name if item.cost_item else item.assembly.name)
                table_data.append([
                    "",
                    Paragraph(item_name, styles["Normal"]),
                    str(item.quantity),
                    item.unit,
                    f"${item.unit_price:,.2f}",
                    f"${item.line_total:,.2f}",
                ])

            # Section subtotal
            table_data.append([
                "", "", "", "",
                Paragraph("<b>Section Total:</b>", styles["Normal"]),
                Paragraph(f"<b>${section.subtotal:,.2f}</b>", styles["Normal"]),
            ])

        # Totals
        table_data.append(["", "", "", "", "", ""])
        table_data.append([
            "", "", "", "",
            Paragraph("<b>Subtotal:</b>", styles["Normal"]),
            Paragraph(f"<b>${estimate.subtotal:,.2f}</b>", styles["Normal"]),
        ])
        table_data.append([
            "", "", "", "",
            Paragraph(f"<b>Tax ({estimate.tax_rate}%):</b>", styles["Normal"]),
            Paragraph(f"<b>${estimate.tax_amount:,.2f}</b>", styles["Normal"]),
        ])
        table_data.append([
            "", "", "", "",
            Paragraph("<b>TOTAL:</b>", styles["Normal"]),
            Paragraph(f"<b>${estimate.total:,.2f}</b>", styles["Normal"]),
        ])

        table = Table(table_data, colWidths=[0.5*inch, 3*inch, 0.6*inch, 0.6*inch, 1*inch, 1*inch])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 12),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
            ("GRID", (0, 0), (-1, -4), 1, colors.grey),
            ("LINEBELOW", (4, -3), (-1, -1), 2, colors.black),
        ]))

        story.append(table)
        story.append(Spacer(1, 0.5*inch))

        # Terms and conditions
        if proposal.terms_and_conditions:
            story.append(Paragraph("<b>Terms and Conditions</b>", heading_style))
            story.append(Spacer(1, 0.1*inch))
            story.append(Paragraph(proposal.terms_and_conditions, styles["Normal"]))
            story.append(Spacer(1, 0.3*inch))

        # Signature section
        if proposal.template and proposal.template.signature_instructions:
            story.append(Paragraph(proposal.template.signature_instructions, styles["Normal"]))
        else:
            story.append(Paragraph("Please sign below to accept this proposal:", styles["Normal"]))

        story.append(Spacer(1, 0.5*inch))
        sig_table_data = [
            ["Signature: _______________________________", "Date: _______________________"],
            ["Name: ____________________________________", ""],
        ]
        sig_table = Table(sig_table_data, colWidths=[4*inch, 2*inch])
        story.append(sig_table)

        # Footer
        if proposal.template and proposal.template.footer_text:
            story.append(Spacer(1, 0.5*inch))
            footer_style = ParagraphStyle(
                "Footer",
                parent=styles["Normal"],
                fontSize=8,
                textColor=colors.grey,
                alignment=TA_CENTER,
            )
            story.append(Paragraph(proposal.template.footer_text, footer_style))

        # Build PDF
        doc.build(story)
        buffer.seek(0)

        return buffer


class AssemblyService:
    """Service for assembly operations."""

    @staticmethod
    def insert_assembly_into_estimate(assembly, section, user, quantity=1):
        """Insert assembly as line item into estimate section."""
        from .models import EstimateLineItem

        with transaction.atomic():
            # Get max sort_order in section
            max_order = section.line_items.aggregate(
                max_order=models.Max("sort_order")
            )["max_order"] or 0

            # Create line item for assembly
            assembly_line = EstimateLineItem.objects.create(
                organization=section.organization,
                section=section,
                assembly=assembly,
                description=assembly.name,
                quantity=quantity,
                unit="EA",
                unit_cost=assembly.total_cost,
                unit_price=assembly.total_price,
                is_taxable=True,
                sort_order=max_order + 1,
            )

            # Recalculate estimate totals
            EstimateCalculationService.calculate_estimate_totals(section.estimate)

            return assembly_line

    @staticmethod
    def calculate_assembly_totals(assembly):
        """Recalculate assembly totals (convenience wrapper)."""
        return EstimateCalculationService.calculate_assembly_totals(assembly)
